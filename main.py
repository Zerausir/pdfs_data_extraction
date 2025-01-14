import os
import re
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill
from openpyxl.utils import get_column_letter
from environs import Env, EnvError

# Inicializar environs
env = Env()
env.read_env()


class PDFExtractor:
    def extract_text_from_pdf(self, pdf_path):
        """
        Extrae el texto completo de un archivo PDF.
        Args:
            pdf_path (str): Ruta al archivo PDF.
        Returns:
            str: Texto extraído del PDF.
        """
        try:
            reader = PdfReader(pdf_path)
            text = "".join(page.extract_text() for page in reader.pages)
            return text
        except Exception as e:
            print(f"Error leyendo el PDF {pdf_path}: {e}")
            return ""

    def extract_field_with_regex(self, text, pattern):
        """
        Extrae un campo específico utilizando expresiones regulares.
        Args:
            text (str): Texto donde buscar la respuesta.
            pattern (str): Patrón regex para extraer el campo.
        Returns:
            str: Resultado encontrado o "No encontrado".
        """
        match = re.search(pattern, text)
        return match.group(1).strip() if match else "No encontrado"

    def process_pdf(self, pdf_path):
        """
        Procesa un PDF para extraer los campos requeridos.
        Args:
            pdf_path (str): Ruta al PDF.
        Returns:
            list: Lista de diccionarios con los datos extraídos.
        """
        text = self.extract_text_from_pdf(pdf_path)
        if not text:
            return []

        # Extraer campos con regex
        doc_name = self.extract_field_with_regex(text, r"(?:Oficio|Memorando) Nro\.\s*(.+?)\n")
        date_match = self.extract_field_with_regex(
            text, r"(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})"
        )
        if date_match != "No encontrado":
            date_parts = re.findall(r"\d+|\w+", date_match)
            months = {"enero": "01", "febrero": "02", "marzo": "03", "abril": "04", "mayo": "05", "junio": "06",
                      "julio": "07", "agosto": "08", "septiembre": "09", "octubre": "10", "noviembre": "11",
                      "diciembre": "12"}
            date = f"{date_parts[0]}-{months[date_parts[2].lower()]}-{date_parts[4]}"
        else:
            date = "No encontrada"

        subject = self.extract_field_with_regex(text, r"(?:Asunto:|ASUNTO:)\s*(.+?)\n")

        references_match = re.search(r"Referencias:\s*(.*?)(?:\n|$)", text, re.DOTALL)
        references = []
        if references_match:
            references_text = references_match.group(1)
            references = [ref.strip() for ref in references_text.split("- ") if ref.strip()]
        references = ", ".join(references) if references else "No encontradas"

        annexes = []
        annexes_section_match = re.search(r"Anexos:\s*(.*?)(?:\n\n|$)", text, re.DOTALL)
        if annexes_section_match:
            annexes_section = annexes_section_match.group(1)
            annexes = re.findall(r"^-\s*(.+)$", annexes_section, re.MULTILINE)

        rows = []
        for annex in annexes:
            rows.append({
                "Nombre": doc_name,
                "Fecha": date,
                "Asunto": subject,
                "Anexo": annex,
                "Referencias": references,
            })

        if not rows:
            rows.append({
                "Nombre": doc_name,
                "Fecha": date,
                "Asunto": subject,
                "Anexo": "No encontrado",
                "Referencias": references,
            })

        return rows

    def process_directory(self, directory):
        """
        Procesa todos los PDFs en un directorio.
        Args:
            directory (str): Ruta al directorio con archivos PDF.
        Returns:
            pd.DataFrame: DataFrame con la información extraída.
        """
        data = []
        for filename in os.listdir(directory):
            if filename.lower().endswith(".pdf"):
                pdf_path = os.path.join(directory, filename)
                data.extend(self.process_pdf(pdf_path))
        return pd.DataFrame(data)


def save_to_excel_with_style(df, output_file):
    """
    Guarda un DataFrame a Excel con filtros y estilos.
    Args:
        df (pd.DataFrame): DataFrame a guardar.
        output_file (str): Ruta del archivo de salida.
    """
    df.to_excel(output_file, index=False, engine="openpyxl")
    wb = load_workbook(output_file)
    ws = wb.active

    # Agregar filtros al encabezado
    ws.auto_filter.ref = ws.dimensions

    # Ajustar ancho de columnas
    for col in ws.columns:
        column_letter = get_column_letter(col[0].column)
        if col[0].value == "Fecha":
            ws.column_dimensions[column_letter].width = 14.00  # Ancho específico para la columna "Fecha"
        else:
            ws.column_dimensions[column_letter].width = 35.00  # Ancho general para las demás columnas

    # Inmovilizar la fila superior
    ws.freeze_panes = "A2"

    # Aplicar estilo predeterminado
    header_style = NamedStyle(name="header_style", font=Font(bold=True), fill=PatternFill("solid", fgColor="D9EAD3"))
    for cell in ws[1]:
        cell.style = header_style

    wb.save(output_file)


def verify_environment_variables():
    """Verifica que las variables de entorno necesarias estén disponibles"""
    required_vars = ['SERVER_ROUTE', 'DOWNLOAD_ROUTE']
    missing_vars = []

    for var in required_vars:
        try:
            env.str(var)
        except EnvError:
            missing_vars.append(var)

    if missing_vars:
        raise EnvironmentError(f"Faltan las siguientes variables de entorno: {', '.join(missing_vars)}")


# Directorio con los archivos PDF
directory = env.str('SERVER_ROUTE')
output_file = os.path.join(env.str('DOWNLOAD_ROUTE'), 'datos_pdfs.xlsx')

# Inicializar extractor y procesar
extractor = PDFExtractor()
df = extractor.process_directory(directory)

# Exportar resultados a Excel con estilo
save_to_excel_with_style(df, output_file)
print(f"Datos extraídos y guardados en {output_file}")
